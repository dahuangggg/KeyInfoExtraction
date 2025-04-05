import os
import json
import pandas as pd
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
from .file_utils import load_json

def save_excel(data, file_path):
    """
    将树状结构数据保存为Excel文件，所有物理状态组放在一个sheet中
    
    参数:
        data: 树状结构数据
        file_path: 输出Excel文件路径
    """
    # 创建一个Excel写入器
    writer = pd.ExcelWriter(file_path, engine='openpyxl')
    
    # 标记是否添加了任何数据
    data_added = False
    
    # 处理树状结构数据
    if "元器件物理状态分析" in data:
        physical_state_analysis = data["元器件物理状态分析"]
        
        # 收集所有物理状态项
        all_rows = []
        
        # 为每个物理状态组添加行
        for group_data in physical_state_analysis:
            # 获取物理状态组名称
            group_name = group_data.get("物理状态组", "未知组")
            
            # 提取物理状态项数据
            physical_state_items = group_data.get("物理状态项", [])
            
            if physical_state_items:
                # 记录该组的行数
                group_start_row = len(all_rows)
                
                # 将物理状态项转换为行数据
                for state in physical_state_items:
                    # 获取物理状态名称和典型物理状态值
                    state_name = state.get("物理状态名称", "")
                    state_value = state.get("典型物理状态值", "")
                    prohibit_info = state.get("禁限用信息", "")
                    test_comment = state.get("测试评语", "")
                    
                    # 处理典型物理状态值可能是字典或列表的情况
                    if isinstance(state_value, dict):
                        # 如果是字典，为每个键值对创建一行
                        first_row = True
                        for key, value in state_value.items():
                            all_rows.append({
                                "物理状态组": group_name if first_row else "",
                                "物理状态名称": state_name if first_row else "",
                                "典型物理状态值": f"{key}: {value}",
                                "禁限用信息": prohibit_info if first_row else "",
                                "测试评语": test_comment if first_row else ""
                            })
                            first_row = False
                    elif isinstance(state_value, list):
                        # 如果是列表，为每个值创建一行
                        for i, value in enumerate(state_value):
                            all_rows.append({
                                "物理状态组": group_name if i == 0 else "",
                                "物理状态名称": state_name if i == 0 else "",
                                "典型物理状态值": value,
                                "禁限用信息": prohibit_info if i == 0 else "",
                                "测试评语": test_comment if i == 0 else ""
                            })
                    else:
                        # 如果不是字典或列表，直接添加一行
                        all_rows.append({
                            "物理状态组": group_name,
                            "物理状态名称": state_name,
                            "典型物理状态值": state_value,
                            "禁限用信息": prohibit_info,
                            "测试评语": test_comment
                        })
        
        # 如果收集到了数据，创建单个工作表
        if all_rows:
            sheet_name = "物理状态分析"
            df = pd.DataFrame(all_rows)
            df.to_excel(writer, sheet_name=sheet_name, index=False)
            data_added = True
            
            # 获取工作簿和工作表对象，用于设置格式
            workbook = writer.book
            worksheet = writer.sheets[sheet_name]
            
            # 设置列宽
            for col_idx, column in enumerate(df.columns):
                column_width = max(df[column].astype(str).map(len).max(), len(column)) + 4
                worksheet.column_dimensions[get_column_letter(col_idx + 1)].width = min(column_width, 40)
            
            # 合并相同物理状态组的单元格和相同物理状态名称的单元格
            current_group = None
            group_start_row = 2  # Excel行从1开始，标题行是1，数据从2开始
            
            current_state = None
            state_start_row = 2
            
            # 对所有行进行处理
            for i, row in enumerate(all_rows):
                row_idx = i + 2  # 加2是因为Excel的索引从1开始，而且有标题行
                
                # 处理物理状态组的合并
                if row["物理状态组"]:  # 如果有物理状态组名称
                    if current_group and row_idx - group_start_row > 1:
                        # 合并之前的物理状态组单元格
                        worksheet.merge_cells(f"A{group_start_row}:A{row_idx-1}")
                    
                    current_group = row["物理状态组"]
                    group_start_row = row_idx
                
                # 处理物理状态名称的合并
                if row["物理状态名称"]:  # 如果有物理状态名称
                    if current_state and row_idx - state_start_row > 1:
                        # 合并之前的物理状态名称单元格
                        worksheet.merge_cells(f"B{state_start_row}:B{row_idx-1}")
                    
                    current_state = row["物理状态名称"]
                    state_start_row = row_idx
            
            # 处理最后一组物理状态组和物理状态名称的合并
            if current_group and len(all_rows) + 2 - group_start_row > 1:
                worksheet.merge_cells(f"A{group_start_row}:A{len(all_rows)+1}")
            
            if current_state and len(all_rows) + 2 - state_start_row > 1:
                worksheet.merge_cells(f"B{state_start_row}:B{len(all_rows)+1}")
            
            # 设置所有单元格内容居中显示
            center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            for row in worksheet.iter_rows(min_row=1, max_row=len(all_rows) + 1, min_col=1, max_col=len(df.columns)):
                for cell in row:
                    cell.alignment = center_alignment
    
    # 如果没有添加任何数据，创建一个默认工作表
    if not data_added:
        # 尝试从数据中提取一些信息
        info_data = {}
        
        # 检查是否有器件信息
        if "器件信息" in data:
            for key, value in data["器件信息"].items():
                info_data[key] = [value]
        
        # 如果没有找到任何信息，添加一个空行
        if not info_data:
            info_data = {"信息": ["没有找到可用的物理状态分析数据"]}
        
        # 创建DataFrame并保存到默认工作表
        df = pd.DataFrame(info_data)
        df.to_excel(writer, sheet_name="基本信息", index=False)
    
    # 保存Excel文件
    writer.close()
    print(f"数据已保存到Excel文件: {file_path}")

def json_to_excel(json_file_path, excel_file_path=None):
    """
    将JSON文件转换为Excel文件
    
    参数:
        json_file_path: JSON文件路径
        excel_file_path: 输出Excel文件路径，如果为None则使用相同的文件名但扩展名为.xlsx
    """
    # 如果未指定Excel文件路径，则使用相同的文件名但扩展名为.xlsx
    if excel_file_path is None:
        excel_file_path = os.path.splitext(json_file_path)[0] + ".xlsx"
    
    # 读取JSON文件
    data = load_json(json_file_path)
    
    # 保存为Excel文件
    if data:
        save_excel(data, excel_file_path)
        return excel_file_path
    return None

# 测试代码
if __name__ == "__main__":
    import sys
    from .file_utils import save_json
    
    # 检查是否提供了JSON文件路径
    if len(sys.argv) > 1:
        json_file_path = sys.argv[1]
        excel_file_path = json_to_excel(json_file_path)
        print(f"已将 {json_file_path} 转换为 {excel_file_path}")
    else:
        # 测试数据
        test_data = {
            "元器件物理状态分析树状结构": {
                "标识部分": {
                    "物理状态组": [
                        {
                            "物理状态名称": "器件型号",
                            "典型物理状态值": "XC9536",
                            "禁限用信息": "无",
                            "测试评语": "清晰可见"
                        },
                        {
                            "物理状态名称": "生产批次",
                            "典型物理状态值": "2023-A",
                            "禁限用信息": "无",
                            "测试评语": "标识完整"
                        }
                    ]
                },
                "封装结构": {
                    "物理状态组": [
                        {
                            "物理状态名称": "封装类型",
                            "典型物理状态值": "PLCC44",
                            "禁限用信息": "无",
                            "测试评语": "符合标准"
                        },
                        {
                            "物理状态名称": "封装材料",
                            "典型物理状态值": ["环氧树脂", "铜引线框架", "镀金层"],
                            "禁限用信息": "无",
                            "测试评语": "质量良好"
                        },
                        {
                            "物理状态名称": "引脚状态",
                            "典型物理状态值": ["完好", "无弯曲", "无氧化"],
                            "禁限用信息": "无",
                            "测试评语": "引脚状态良好"
                        }
                    ]
                }
            }
        }
        
        # 保存测试数据为JSON
        test_json_path = "test_tree_output.json"
        save_json(test_data, test_json_path)
        
        # 转换为Excel
        test_excel_path = json_to_excel(test_json_path)
        print(f"已创建测试Excel文件: {test_excel_path}") 